from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.filters import CommandStart, Command, Filter, StateFilter
from aiogram.fsm.context import FSMContext

import os
from openpyxl import Workbook
from datetime import datetime

from app.database.requests import (
    get_all_points, get_all_zones, get_all_regions, 
    get_points_by_zone, get_zones_by_region,
    get_report_data, get_user_by_tg_id, 
    add_point, add_region, add_zone, 
    get_point_by_id,get_region_by_id,get_zone_by_id, 
    add_shipment, get_user_by_point_id,
    update_bags_count, get_all_requests_sorted,
    get_all_shipments_sorted, get_combined_data_sorted, delete_point_and_related_data
)
from app.states import Reports, ShipmentStates, CreatePoint
from app.keyboards import (
    report_keyboard, admin_keyboard, driver_keyboard,
    cancel_keyboard, confirm_keyboard, 
    get_cancel_keyboard, get_category_keyboard, get_confirmation_keyboard, 
    get_main_materials_keyboard, get_other_materials_keyboard)

ADMIN_IDS = [753755508, 1582399282, 7854337092, 7854337092, 7363212828, 6700699811]  # ID администраторов (оставляем без изменений)

admin = Router()

class Admin(Filter):
    async def __call__(self, message: Message) -> bool:
        return message.from_user.id in ADMIN_IDS

@admin.message(Admin(), Command("A1"))
async def admin_start(message: Message):
    """Главное меню админ-панели"""
    await message.answer(
        "ยินดีต้อนรับสู่แผงควบคุมผู้ดูแล กรุณาเลือกการดำเนินการ:",  # "Добро пожаловать в админ-панель. Выберите действие:"
        reply_markup=admin_keyboard()
    )

@admin.callback_query(Admin(), F.data == 'report')
async def cmd_report(callback: CallbackQuery, state: FSMContext):
    """Меню отчетов"""
    await callback.message.answer('กรุณาเลือกรูปแบบรายงาน:', reply_markup=report_keyboard())  # "Выберите тип отчета:"
    await state.set_state(Reports.report_type)

@admin.callback_query(Admin(), Reports.report_type, F.data == "report_zone")
async def process_zones_report(callback: CallbackQuery):
    """Отчет по зонам"""
    zones = await get_all_zones()
    
    if not zones:
        await callback.message.answer('ไม่มีข้อมูลโซน')  # "Нет данных по зонам."
        return
    
    report_messages = ["รายงานตามโซน:\n\n"]  # "Отчет по зонам:"
    for zone in zones:
        points = await get_points_by_zone(zone.zone_id)
        total_bags = sum(p.bags_count for p in points) if points else 0
        
        zone_info = (
            f"Зона ID: {zone.zone_id}\n"
            f"Регион: {zone.region_id}\n"
            f"Точек: {len(points)}\n"
            f"Всего мешков: {total_bags}\n\n"
        )
        
        # Если добавление новой информации превысит лимит - создаем новое сообщение
        if len(report_messages[-1]) + len(zone_info) > 4000:
            report_messages.append(zone_info)
        else:
            report_messages[-1] += zone_info
    
    for msg in report_messages:
        await callback.message.answer(msg)
    
    await callback.message.answer('กรุณาเลือกการดำเนินการ:', reply_markup=admin_keyboard())  # "Выберите действие:"

@admin.callback_query(Admin(), Reports.report_type, F.data == "report_region")
async def process_regions_report(callback: CallbackQuery):
    """Отчет по регионам"""
    regions = await get_all_regions()
    
    if not regions:
        await callback.message.answer('ไม่มีข้อมูลภูมิภาค')  # "Нет данных по регионам."
        return
    
    report_messages = ["รายงานตามภูมิภาค:\n\n"]  # "Отчет по регионам:"
    for region in regions:
        zones = await get_zones_by_region(region.region_id)
        zone_count = len(zones)
        point_count = 0
        total_bags = 0
        
        for zone in zones:
            points = await get_points_by_zone(zone.zone_id)
            point_count += len(points)
            total_bags += sum(p.bags_count for p in points)
        
        region_info = (
            f"ภูมิภาค ID: {region.region_id}\n"  # "Регион ID:"
            f"จำนวนโซน: {zone_count}\n"  # "Зон:"
            f"จำนวนจุด: {point_count}\n"  # "Точек:"
            f"จำนวนถุงทั้งหมด: {total_bags}\n\n"  # "Всего мешков:"
        )
        
        if len(report_messages[-1]) + len(region_info) > 4000:
            report_messages.append(region_info)
        else:
            report_messages[-1] += region_info
    
    for msg in report_messages:
        await callback.message.answer(msg)
    
    await callback.message.answer('กรุณาเลือกการดำเนินการ:', reply_markup=admin_keyboard())  # "Выберите действие:"

@admin.callback_query(Admin(), Reports.report_type, F.data == "report_region_detail")
async def ask_region_id(callback: CallbackQuery, state: FSMContext):
    """Запрос ID региона для детального отчета"""
    await callback.message.answer("กรุณากรอก ID ภูมิภาคสำหรับรายงาน:")  # "Введите ID региона для отчета:"
    await state.set_state(Reports.waiting_region_id)

@admin.message(Admin(), Reports.waiting_region_id)
async def generate_region_detail_report(message: Message, state: FSMContext):
    """Генерация детального отчета по региону"""
    try:
        region_id = int(message.text.strip())
    except ValueError:
        await message.answer("กรุณากรอก ID ภูมิภาคเป็นตัวเลข")  # "Пожалуйста, введите числовой ID региона."
        return
    
    # Получаем данные по региону
    region = await get_region_by_id(region_id)
    if not region:
        await message.answer(f"ไม่พบภูมิภาค ID {region_id}")  # f"Регион с ID {region_id} не найден."
        await state.clear()
        return
    
    # Получаем все зоны в регионе
    zones = await get_zones_by_region(region_id)
    if not zones:
        await message.answer(f"ไม่มีโซนในภูมิภาค {region_id}")  # f"В регионе {region_id} нет зон."
        await state.clear()
        return
    
    # Считаем общую статистику по региону
    total_points = 0
    total_bags = 0
    zones_data = []
    
    for zone in zones:
        points = await get_points_by_zone(zone.zone_id)
        zone_points = len(points)
        zone_bags = sum(p.bags_count for p in points) if points else 0
        
        total_points += zone_points
        total_bags += zone_bags
        
        if zone_points > 0:
            zones_data.append((zone.zone_id, zone_points, zone_bags))
    
    # Формируем сообщения с учетом лимита
    header = (
        f"📊 รายงานภูมิภาค {region_id}:\n"  # f"📊 Отчет по региону {region_id}:"
        f"จำนวนจุดทั้งหมด: {total_points}\n"  # f"Всего точек: {total_points}"
        f"จำนวนถุงทั้งหมด: {total_bags}\n\n"  # f"Всего мешков готово: {total_bags}"
        f"รายละเอียดโซน:\n"  # f"Детали по зонам:"
    )
    
    report_messages = [header]
    
    # Добавляем информацию по зонам
    for zone_id, zone_points, zone_bags in zones_data:
        zone_info = f"▪️ โซน {zone_id} - {zone_points} จุด ({zone_bags} ถุง)\n"  # f"▪️ Зона {zone_id} - {zone_points} точек ({zone_bags} мешков)"
        
        if len(report_messages[-1]) + len(zone_info) > 4000:
            report_messages.append(zone_info)
        else:
            report_messages[-1] += zone_info
    
    for msg in report_messages:
        await message.answer(msg)
    
    await message.answer('กรุณาเลือกการดำเนินการ:', reply_markup=admin_keyboard())  # "Выберите действие:"
    await state.clear()

@admin.callback_query(Admin(), F.data == "generate_log_report")
async def generate_log_report(callback: CallbackQuery):
    """Генерация Excel-отчета с тремя листами: заявки, отгрузки и объединенные данные"""
    requests = await get_all_requests_sorted()
    shipments = await get_all_shipments_sorted()
    combined = await get_combined_data_sorted()
    
    if not requests and not shipments:
        await callback.message.answer("Нет данных для формирования отчета.")
        return
    
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # 1. Лист с заявками (тайский + английский)
    if requests:
        ws_requests = wb.create_sheet("Requests")
        headers_requests = [
            "Date", 
            "Point ID", 
            "User ID", 
            "Activity Type",
            "PET Bags", 
            "Aluminum Bags", 
            "Glass Bags", 
            "Other Bags",
            "Total Quantity"
        ]
        ws_requests.append(headers_requests)
        
        for req in requests:
            ws_requests.append([
                req.timestamp.strftime('%Y-%m-%d %H:%M'),
                req.point_id,
                req.user_id,
                req.activity,
                req.pet_bag or 0,
                req.aluminum_bag or 0,
                req.glass_bag or 0,
                req.other or 0,
                (req.pet_bag or 0) + (req.aluminum_bag or 0) + 
                (req.glass_bag or 0) + (req.other or 0)
            ])
    
    # 2. Лист с отгрузками (тайский + английский)
    if shipments:
        ws_shipments = wb.create_sheet("Shipments")
        headers_shipments = [
            "Date", 
            "Point ID", 
            "Driver ID", 
            "Total Payment",
            # Основные материалы (категория 1)
            "PET kg", "PET Price", "PET Total",
            "Paper kg", "Paper Price", "Paper Total",
            "Aluminum kg", "Aluminum Price", "Aluminum Total",
            "Glass kg", "Glass Price", "Glass Total",
            "Small Beer Box kg", "Small Beer Box Price", "Small Beer Box Total",
            "Large Beer Box kg", "Large Beer Box Price", "Large Beer Box Total",
            "Mixed Beer Box kg", "Mixed Beer Box Price", "Mixed Beer Box Total",
            # Другие материалы (категория 2)
            "Oil kg", "Oil Price", "Oil Total",
            "Colored Plastic kg", "Colored Plastic Price", "Colored Plastic Total",
            "Iron kg", "Iron Price", "Iron Total",
            "Plastic Bag kg", "Plastic Bag Price", "Plastic Bag Total",
            "Mix kg", "Mix Price", "Mix Total",
            "Other kg", "Other Price", "Other Total"
        ]
        ws_shipments.append(headers_shipments)
        
        for ship in shipments:
            ws_shipments.append([
                ship.timestamp.strftime('%Y-%m-%d %H:%M'),
                ship.point_id,
                ship.user_id,
                ship.total_pay,
                # Основные материалы
                ship.pet_kg, ship.pet_price, ship.pet_total,
                ship.paper_kg, ship.paper_price, ship.paper_total,
                ship.alum_kg, ship.alum_price, ship.alum_total,
                ship.glass_kg, ship.glass_price, ship.glass_total,
                ship.small_beer_box_kg, ship.small_beer_box_price, ship.small_beer_box_total,
                ship.large_beer_box_kg, ship.large_beer_box_price, ship.large_beer_box_total,
                ship.mixed_beer_box_kg, ship.mixed_beer_box_price, ship.mixed_beer_box_total,
                # Другие материалы
                ship.oil_kg, ship.oil_price, ship.oil_total,
                ship.colored_plastic_kg, ship.colored_plastic_price, ship.colored_plastic_total,
                ship.iron_kg, ship.iron_price, ship.iron_total,
                ship.plastic_bag_kg, ship.plastic_bag_price, ship.plastic_bag_total,
                ship.mix_kg, ship.mix_price, ship.mix_total,
                ship.other_kg, ship.other_price, ship.other_total
            ])
    
    # 3. Общий лист (тайский + английский)
    if combined:
        ws_combined = wb.create_sheet("All Data")
        headers_combined = [
            "Type", 
            "Date", 
            "Point ID", 
            "User ID",
            # Основные материалы
            "PET (Bags/kg)", "PET Price", "PET Total",
            "Paper kg", "Paper Price", "Paper Total",
            "Aluminum (Bags/kg)", "Aluminum Price", "Aluminum Total",
            "Glass (Bags/kg)", "Glass Price", "Glass Total",
            "Small Beer Box kg", "Small Beer Box Price", "Small Beer Box Total",
            "Large Beer Box kg", "Large Beer Box Price", "Large Beer Box Total",
            "Mixed Beer Box kg", "Mixed Beer Box Price", "Mixed Beer Box Total",
            # Другие материалы
            "Oil kg", "Oil Price", "Oil Total",
            "Colored Plastic kg", "Colored Plastic Price", "Colored Plastic Total",
            "Iron kg", "Iron Price", "Iron Total",
            "Plastic Bag kg", "Plastic Bag Price", "Plastic Bag Total",
            "Mix kg", "Mix Price", "Mix Total",
            "Other kg", "Other Price", "Other Total",
            # Итоги
            "Total Amount", 
            "Activity Type (for requests)"
        ]
        ws_combined.append(headers_combined)
        
        for item in combined:
            if item["type"] == "request":
                row = [
                    "Request",
                    item["timestamp"].strftime('%Y-%m-%d %H:%M'),
                    item["point_id"],
                    item["user_id"],
                    item["pet"] or 0, "", "",  # PET (Bags)
                    "", "", "",  # Paper
                    item["aluminum"] or 0, "", "",  # Aluminum (Bags)
                    item["glass"] or 0, "", "",  # Glass (Bags)
                    "", "", "",  # Small Beer Box
                    "", "", "",  # Large Beer Box
                    "", "", "",  # Mixed Beer Box
                    "", "", "",  # Oil
                    "", "", "",  # Colored Plastic
                    "", "", "",  # Iron
                    "", "", "",  # Plastic Bag
                    "", "", "",  # Mix
                    item["other"] or 0, "", "",  # Other
                    "",  # Total Amount
                    item["activity"]  # Activity Type
                ]
            else:
                row = [
                    "Shipment",
                    item["timestamp"].strftime('%Y-%m-%d %H:%M'),
                    item["point_id"],
                    item["user_id"],
                    # PET
                    item["pet_kg"], item["pet_price"], item["pet_total"],
                    # Paper
                    item["paper_kg"], item["paper_price"], item["paper_total"],
                    # Aluminum
                    item["alum_kg"], item["alum_price"], item["alum_total"],
                    # Glass
                    item["glass_kg"], item["glass_price"], item["glass_total"],
                    # Small Beer Box
                    item["small_beer_box_kg"], item["small_beer_box_price"], item["small_beer_box_total"],
                    # Large Beer Box
                    item["large_beer_box_kg"], item["large_beer_box_price"], item["large_beer_box_total"],
                    # Mixed Beer Box
                    item["mixed_beer_box_kg"], item["mixed_beer_box_price"], item["mixed_beer_box_total"],
                    # Oil
                    item["oil_kg"], item["oil_price"], item["oil_total"],
                    # Colored Plastic
                    item["colored_plastic_kg"], item["colored_plastic_price"], item["colored_plastic_total"],
                    # Iron
                    item["iron_kg"], item["iron_price"], item["iron_total"],
                    # Plastic Bag
                    item["plastic_bag_kg"], item["plastic_bag_price"], item["plastic_bag_total"],
                    # Mix
                    item["mix_kg"], item["mix_price"], item["mix_total"],
                    # Other
                    item["other_kg"], item["other_price"], item["other_total"],
                    # Total
                    item["total_pay"],
                    ""  # Activity Type (empty for shipments)
                ]
            ws_combined.append(row)
    
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    await callback.message.answer_document(FSInputFile(filename), reply_markup=admin_keyboard())
    os.remove(filename)

# Основное меню водителя
@admin.message(Command('D1'))
async def cmd_driver(message: Message):
    await message.answer("กรุณาเลือกการดำเนินการ:", reply_markup=driver_keyboard())  # "Выберите действие:"

@admin.callback_query(F.data == "add_shipment")
async def process_add_shipment(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer('กรุณากรอก ID จุด:')  # "Введите ID точки:"
    await state.set_state(ShipmentStates.point_id)
    await callback.answer()

@admin.message(ShipmentStates.point_id)
async def process_point_id(message: Message, state: FSMContext):
    try:
        point_id = int(message.text)
        # Проверяем существование точки
        point = await get_point_by_id(point_id)
        if not point:
            await message.answer("ข้อผิดพลาด: ไม่พบจุดนี้ในระบบ กรุณากรอกใหม่")  # "Ошибка: Точка не найдена в системе. Пожалуйста, введите ID точки заново."
            return
            
        await state.update_data(point_id=point_id)
        await message.answer('กรุณาเลือกประเภท:', reply_markup=get_category_keyboard()) #Пожалуйста, выберите тип
    except ValueError:
        await message.answer("ข้อผิดพลาด: ID จุดต้องเป็นตัวเลขเต็ม กรุณากรอกใหม่") #Ошибка: идентификатор точки должен быть целым числом. пожалуйста, введите новый.

# Обработчики выбора категорий
@admin.callback_query(F.data == "category_main")
async def select_main_materials(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("กรุณาเลือกวัสดุ:", reply_markup=get_main_materials_keyboard())  # "Выберите материал:"
    await callback.answer()

@admin.callback_query(F.data == "category_other")
async def select_other_materials(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("กรุณาเลือกวัสดุ:", reply_markup=get_other_materials_keyboard())  # "Выберите материал:"
    await callback.answer()

@admin.callback_query(F.data == "back_to_categories")
async def back_to_categories(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("กรุณาเลือกประเภท:", reply_markup=get_category_keyboard())  # "Выберите категорию:"
    await callback.answer()

# Обработчики для основных материалов
@admin.callback_query(F.data == "material_pet")
async def process_pet_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Plastic PET (กก.):")  # "Введите вес Plastic PET (кг):"
    await state.set_state(ShipmentStates.pet_kg)
    await callback.answer()

@admin.message(ShipmentStates.pet_kg)
async def process_pet_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        pet_kg = float(text)
        if pet_kg < 0:
            raise ValueError
        await state.update_data(pet_kg=pet_kg)
        
        if pet_kg == 0:
            await state.update_data(pet_price=0.0)
            await message.answer("น้ำหนัก Plastic PET เป็น 0", reply_markup=get_category_keyboard())  # "Вес Plastic PET равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Plastic PET:")  # "Введите цену за кг Plastic PET:"
            await state.set_state(ShipmentStates.pet_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.pet_price)
async def process_pet_price(message: Message, state: FSMContext):
    try:
        pet_price = float(message.text)
        if pet_price < 0:
            raise ValueError
        await state.update_data(pet_price=pet_price)
        await message.answer("บันทึกข้อมูล Plastic PET เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Plastic PET сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Paper
@admin.callback_query(F.data == "material_paper")
async def process_paper_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Paper (กก.):")  # "Введите вес Paper (кг):"
    await state.set_state(ShipmentStates.paper_kg)
    await callback.answer()

@admin.message(ShipmentStates.paper_kg)
async def process_paper_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        paper_kg = float(text)
        if paper_kg < 0:
            raise ValueError
        await state.update_data(paper_kg=paper_kg)
        
        if paper_kg == 0:
            await state.update_data(paper_price=0.0)
            await message.answer("น้ำหนัก Paper เป็น 0", reply_markup=get_category_keyboard())  # "Вес Paper равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Paper:")  # "Введите цену за кг Paper:"
            await state.set_state(ShipmentStates.paper_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.paper_price)
async def process_paper_price(message: Message, state: FSMContext):
    try:
        paper_price = float(message.text)
        if paper_price < 0:
            raise ValueError
        await state.update_data(paper_price=paper_price)
        await message.answer("บันทึกข้อมูล Paper เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Paper сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Aluminum
@admin.callback_query(F.data == "material_alum")
async def process_alum_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Aluminum (กก.):")  # "Введите вес Aluminum (кг):"
    await state.set_state(ShipmentStates.alum_kg)
    await callback.answer()

@admin.message(ShipmentStates.alum_kg)
async def process_alum_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        alum_kg = float(text)
        if alum_kg < 0:
            raise ValueError
        await state.update_data(alum_kg=alum_kg)
        
        if alum_kg == 0:
            await state.update_data(alum_price=0.0)
            await message.answer("น้ำหนัก Aluminum เป็น 0", reply_markup=get_category_keyboard())  # "Вес Aluminum равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Aluminum:")  # "Введите цену за кг Aluminum:"
            await state.set_state(ShipmentStates.alum_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.alum_price)
async def process_alum_price(message: Message, state: FSMContext):
    try:
        alum_price = float(message.text)
        if alum_price < 0:
            raise ValueError
        await state.update_data(alum_price=alum_price)
        await message.answer("บันทึกข้อมูล Aluminum เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Aluminum сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Glass
@admin.callback_query(F.data == "material_glass")
async def process_glass_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Glass (กก.):")  # "Введите вес Glass (кг):"
    await state.set_state(ShipmentStates.glass_kg)
    await callback.answer()

@admin.message(ShipmentStates.glass_kg)
async def process_glass_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        glass_kg = float(text)
        if glass_kg < 0:
            raise ValueError
        await state.update_data(glass_kg=glass_kg)
        
        if glass_kg == 0:
            await state.update_data(glass_price=0.0)
            await message.answer("น้ำหนัก Glass เป็น 0", reply_markup=get_category_keyboard())  # "Вес Glass равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Glass:")  # "Введите цену за кг Glass:"
            await state.set_state(ShipmentStates.glass_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.glass_price)
async def process_glass_price(message: Message, state: FSMContext):
    try:
        glass_price = float(message.text)
        if glass_price < 0:
            raise ValueError
        await state.update_data(glass_price=glass_price)
        await message.answer("บันทึกข้อมูล Glass เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Glass сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Small Beer Box
@admin.callback_query(F.data == "material_small_beer_box")
async def process_small_beer_box_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Small Beer Box (กก.):")
    await state.set_state(ShipmentStates.small_beer_box_kg)
    await callback.answer()

@admin.message(ShipmentStates.small_beer_box_kg)
async def process_small_beer_box_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        small_beer_box_kg = float(text)
        if small_beer_box_kg < 0:
            raise ValueError
        await state.update_data(small_beer_box_kg=small_beer_box_kg)
        
        if small_beer_box_kg == 0:
            await state.update_data(small_beer_box_price=0.0)
            await message.answer("น้ำหนัก Small Beer Box เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Small Beer Box:")
            await state.set_state(ShipmentStates.small_beer_box_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.small_beer_box_price)
async def process_small_beer_box_price(message: Message, state: FSMContext):
    try:
        small_beer_box_price = float(message.text)
        if small_beer_box_price < 0:
            raise ValueError
        await state.update_data(small_beer_box_price=small_beer_box_price)
        await message.answer("บันทึกข้อมูล Small Beer Box เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Large Beer Box
@admin.callback_query(F.data == "material_large_beer_box")
async def process_large_beer_box_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Large Beer Box (กก.):")
    await state.set_state(ShipmentStates.large_beer_box_kg)
    await callback.answer()

@admin.message(ShipmentStates.large_beer_box_kg)
async def process_large_beer_box_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        large_beer_box_kg = float(text)
        if large_beer_box_kg < 0:
            raise ValueError
        await state.update_data(large_beer_box_kg=large_beer_box_kg)
        
        if large_beer_box_kg == 0:
            await state.update_data(large_beer_box_price=0.0)
            await message.answer("น้ำหนัก Large Beer Box เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Large Beer Box:")
            await state.set_state(ShipmentStates.large_beer_box_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.large_beer_box_price)
async def process_large_beer_box_price(message: Message, state: FSMContext):
    try:
        large_beer_box_price = float(message.text)
        if large_beer_box_price < 0:
            raise ValueError
        await state.update_data(large_beer_box_price=large_beer_box_price)
        await message.answer("บันทึกข้อมูล Large Beer Box เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Mixed Beer Box
@admin.callback_query(F.data == "material_mixed_beer_box")
async def process_mixed_beer_box_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Mixed Beer Box (กก.):")
    await state.set_state(ShipmentStates.mixed_beer_box_kg)
    await callback.answer()

@admin.message(ShipmentStates.mixed_beer_box_kg)
async def process_mixed_beer_box_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        mixed_beer_box_kg = float(text)
        if mixed_beer_box_kg < 0:
            raise ValueError
        await state.update_data(mixed_beer_box_kg=mixed_beer_box_kg)
        
        if mixed_beer_box_kg == 0:
            await state.update_data(mixed_beer_box_price=0.0)
            await message.answer("น้ำหนัก Mixed Beer Box เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Mixed Beer Box:")
            await state.set_state(ShipmentStates.mixed_beer_box_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.mixed_beer_box_price)
async def process_mixed_beer_box_price(message: Message, state: FSMContext):
    try:
        mixed_beer_box_price = float(message.text)
        if mixed_beer_box_price < 0:
            raise ValueError
        await state.update_data(mixed_beer_box_price=mixed_beer_box_price)
        await message.answer("บันทึกข้อมูล Mixed Beer Box เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для других материалов (категория 2)
# Обработчики для Oil
@admin.callback_query(F.data == "material_oil")
async def process_oil_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Oil (กก.):")  # "Введите вес Oil (кг):"
    await state.set_state(ShipmentStates.oil_kg)
    await callback.answer()

@admin.message(ShipmentStates.oil_kg)
async def process_oil_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        oil_kg = float(text)
        if oil_kg < 0:
            raise ValueError
        await state.update_data(oil_kg=oil_kg)
        
        if oil_kg == 0:
            await state.update_data(oil_price=0.0)
            await message.answer("น้ำหนัก Oil เป็น 0", reply_markup=get_category_keyboard())  # "Вес Oil равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Oil:")  # "Введите цену за кг Oil:"
            await state.set_state(ShipmentStates.oil_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.oil_price)
async def process_oil_price(message: Message, state: FSMContext):
    try:
        oil_price = float(message.text)
        if oil_price < 0:
            raise ValueError
        await state.update_data(oil_price=oil_price)
        await message.answer("บันทึกข้อมูล Oil เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Oil сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Colored Plastic
@admin.callback_query(F.data == "material_colored_plastic")
async def process_colored_plastic_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Colored Plastic (กก.):")
    await state.set_state(ShipmentStates.colored_plastic_kg)
    await callback.answer()

@admin.message(ShipmentStates.colored_plastic_kg)
async def process_colored_plastic_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        colored_plastic_kg = float(text)
        if colored_plastic_kg < 0:
            raise ValueError
        await state.update_data(colored_plastic_kg=colored_plastic_kg)
        
        if colored_plastic_kg == 0:
            await state.update_data(colored_plastic_price=0.0)
            await message.answer("น้ำหนัก Colored Plastic เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Colored Plastic:")
            await state.set_state(ShipmentStates.colored_plastic_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.colored_plastic_price)
async def process_colored_plastic_price(message: Message, state: FSMContext):
    try:
        colored_plastic_price = float(message.text)
        if colored_plastic_price < 0:
            raise ValueError
        await state.update_data(colored_plastic_price=colored_plastic_price)
        await message.answer("บันทึกข้อมูล Colored Plastic เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Iron
@admin.callback_query(F.data == "material_iron")
async def process_iron_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Iron (กก.):")
    await state.set_state(ShipmentStates.iron_kg)
    await callback.answer()

@admin.message(ShipmentStates.iron_kg)
async def process_iron_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        iron_kg = float(text)
        if iron_kg < 0:
            raise ValueError
        await state.update_data(iron_kg=iron_kg)
        
        if iron_kg == 0:
            await state.update_data(iron_price=0.0)
            await message.answer("น้ำหนัก Iron เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Iron:")
            await state.set_state(ShipmentStates.iron_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.iron_price)
async def process_iron_price(message: Message, state: FSMContext):
    try:
        iron_price = float(message.text)
        if iron_price < 0:
            raise ValueError
        await state.update_data(iron_price=iron_price)
        await message.answer("บันทึกข้อมูล Iron เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Plastic Bag or Container
@admin.callback_query(F.data == "material_plastic_bag")
async def process_plastic_bag_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Plastic Bag or Container (กก.):")
    await state.set_state(ShipmentStates.plastic_bag_kg)
    await callback.answer()

@admin.message(ShipmentStates.plastic_bag_kg)
async def process_plastic_bag_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        plastic_bag_kg = float(text)
        if plastic_bag_kg < 0:
            raise ValueError
        await state.update_data(plastic_bag_kg=plastic_bag_kg)
        
        if plastic_bag_kg == 0:
            await state.update_data(plastic_bag_price=0.0)
            await message.answer("น้ำหนัก Plastic Bag or Container เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Plastic Bag or Container:")
            await state.set_state(ShipmentStates.plastic_bag_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.plastic_bag_price)
async def process_plastic_bag_price(message: Message, state: FSMContext):
    try:
        plastic_bag_price = float(message.text)
        if plastic_bag_price < 0:
            raise ValueError
        await state.update_data(plastic_bag_price=plastic_bag_price)
        await message.answer("บันทึกข้อมูล Plastic Bag or Container เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Mix
@admin.callback_query(F.data == "material_mix")
async def process_mix_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Mix (กก.):")
    await state.set_state(ShipmentStates.mix_kg)
    await callback.answer()

@admin.message(ShipmentStates.mix_kg)
async def process_mix_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        mix_kg = float(text)
        if mix_kg < 0:
            raise ValueError
        await state.update_data(mix_kg=mix_kg)
        
        if mix_kg == 0:
            await state.update_data(mix_price=0.0)
            await message.answer("น้ำหนัก Mix เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Mix:")
            await state.set_state(ShipmentStates.mix_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.mix_price)
async def process_mix_price(message: Message, state: FSMContext):
    try:
        mix_price = float(message.text)
        if mix_price < 0:
            raise ValueError
        await state.update_data(mix_price=mix_price)
        await message.answer("บันทึกข้อมูล Mix เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Other
@admin.callback_query(F.data == "material_other")
async def process_other_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Other (กก.):")
    await state.set_state(ShipmentStates.other_kg)
    await callback.answer()

@admin.message(ShipmentStates.other_kg)
async def process_other_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        other_kg = float(text)
        if other_kg < 0:
            raise ValueError
        await state.update_data(other_kg=other_kg)
        
        if other_kg == 0:
            await state.update_data(other_price=0.0)
            await message.answer("น้ำหนัก Other เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Other:")
            await state.set_state(ShipmentStates.other_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.other_price)
async def process_other_price(message: Message, state: FSMContext):
    try:
        other_price = float(message.text)
        if other_price < 0:
            raise ValueError
        await state.update_data(other_price=other_price)
        await message.answer("บันทึกข้อมูล Other เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчик завершения ввода
@admin.callback_query(F.data == "finish_shipment")
async def finish_shipment(callback: CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    
    # Формируем сводку по всем введенным данным
    summary = "📋 สรุปการจัดส่ง:\n\n"  # "📋 Сводка по отгрузке:"
    total_weight = 0
    total_cost = 0
    
    # Основные материалы
    if 'pet_kg' in user_data:
        pet_cost = user_data.get('pet_kg', 0) * user_data.get('pet_price', 0)
        summary += f"🔹 Plastic PET: {user_data['pet_kg']} กก., {pet_cost:.2f} บาท\n"
        total_weight += user_data['pet_kg']
        total_cost += pet_cost
    
    if 'paper_kg' in user_data:
        paper_cost = user_data.get('paper_kg', 0) * user_data.get('paper_price', 0)
        summary += f"🔹 Paper: {user_data['paper_kg']} กก., {paper_cost:.2f} บาท\n"
        total_weight += user_data['paper_kg']
        total_cost += paper_cost
    
    if 'alum_kg' in user_data:
        alum_cost = user_data.get('alum_kg', 0) * user_data.get('alum_price', 0)
        summary += f"🔹 Aluminum: {user_data['alum_kg']} กก., {alum_cost:.2f} บาท\n"
        total_weight += user_data['alum_kg']
        total_cost += alum_cost
    
    if 'glass_kg' in user_data:
        glass_cost = user_data.get('glass_kg', 0) * user_data.get('glass_price', 0)
        summary += f"🔹 Glass: {user_data['glass_kg']} กก., {glass_cost:.2f} บาท\n"
        total_weight += user_data['glass_kg']
        total_cost += glass_cost
    
    if 'small_beer_box_kg' in user_data:
        small_beer_box_cost = user_data.get('small_beer_box_kg', 0) * user_data.get('small_beer_box_price', 0)
        summary += f"🔹 Small Beer Box: {user_data['small_beer_box_kg']} กก., {small_beer_box_cost:.2f} บาท\n"
        total_weight += user_data['small_beer_box_kg']
        total_cost += small_beer_box_cost
    
    if 'large_beer_box_kg' in user_data:
        large_beer_box_cost = user_data.get('large_beer_box_kg', 0) * user_data.get('large_beer_box_price', 0)
        summary += f"🔹 Large Beer Box: {user_data['large_beer_box_kg']} กก., {large_beer_box_cost:.2f} บาท\n"
        total_weight += user_data['large_beer_box_kg']
        total_cost += large_beer_box_cost
    
    if 'mixed_beer_box_kg' in user_data:
        mixed_beer_box_cost = user_data.get('mixed_beer_box_kg', 0) * user_data.get('mixed_beer_box_price', 0)
        summary += f"🔹 Mixed Beer Box: {user_data['mixed_beer_box_kg']} กก., {mixed_beer_box_cost:.2f} บาท\n"
        total_weight += user_data['mixed_beer_box_kg']
        total_cost += mixed_beer_box_cost
    
    # Другие материалы
    if 'oil_kg' in user_data:
        oil_cost = user_data.get('oil_kg', 0) * user_data.get('oil_price', 0)
        summary += f"🔸 Oil: {user_data['oil_kg']} กก., {oil_cost:.2f} บาท\n"
        total_weight += user_data['oil_kg']
        total_cost += oil_cost
    
    if 'colored_plastic_kg' in user_data:
        colored_plastic_cost = user_data.get('colored_plastic_kg', 0) * user_data.get('colored_plastic_price', 0)
        summary += f"🔸 Colored Plastic: {user_data['colored_plastic_kg']} กก., {colored_plastic_cost:.2f} บาท\n"
        total_weight += user_data['colored_plastic_kg']
        total_cost += colored_plastic_cost
    
    if 'iron_kg' in user_data:
        iron_cost = user_data.get('iron_kg', 0) * user_data.get('iron_price', 0)
        summary += f"🔸 Iron: {user_data['iron_kg']} กก., {iron_cost:.2f} บาท\n"
        total_weight += user_data['iron_kg']
        total_cost += iron_cost
    
    if 'plastic_bag_kg' in user_data:
        plastic_bag_cost = user_data.get('plastic_bag_kg', 0) * user_data.get('plastic_bag_price', 0)
        summary += f"🔸 Plastic Bag or Container: {user_data['plastic_bag_kg']} กก., {plastic_bag_cost:.2f} บาท\n"
        total_weight += user_data['plastic_bag_kg']
        total_cost += plastic_bag_cost
    
    if 'mix_kg' in user_data:
        mix_cost = user_data.get('mix_kg', 0) * user_data.get('mix_price', 0)
        summary += f"🔸 Mix: {user_data['mix_kg']} กก., {mix_cost:.2f} บาท\n"
        total_weight += user_data['mix_kg']
        total_cost += mix_cost
    
    if 'other_kg' in user_data:
        other_cost = user_data.get('other_kg', 0) * user_data.get('other_price', 0)
        summary += f"🔸 Other: {user_data['other_kg']} กก., {other_cost:.2f} บาท\n"
        total_weight += user_data['other_kg']
        total_cost += other_cost
    
    summary += f"\n💎 รวม: {total_weight:.2f} กก., {total_cost:.2f} บาท"  # "💎 Итого: кг, руб."
    
    await callback.message.edit_text(summary, reply_markup=get_confirmation_keyboard())
    await callback.answer()

# Обработчик подтверждения
@admin.callback_query(F.data == "confirm_shipment")
async def confirm_shipment(callback: CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    
    try:
        user = await get_user_by_tg_id(callback.from_user.id)
        if not user:
            await callback.answer()  # Просто закрываем callback без сообщения
            return  # Выходим из функции, если пользователь не найден
        
        # Рассчитываем общие суммы для каждого материала
        pet_total = user_data.get('pet_kg', 0.0) * user_data.get('pet_price', 0.0)
        paper_total = user_data.get('paper_kg', 0.0) * user_data.get('paper_price', 0.0)
        alum_total = user_data.get('alum_kg', 0.0) * user_data.get('alum_price', 0.0)
        glass_total = user_data.get('glass_kg', 0.0) * user_data.get('glass_price', 0.0)
        small_beer_box_total = user_data.get('small_beer_box_kg', 0.0) * user_data.get('small_beer_box_price', 0.0)
        large_beer_box_total = user_data.get('large_beer_box_kg', 0.0) * user_data.get('large_beer_box_price', 0.0)
        mixed_beer_box_total = user_data.get('mixed_beer_box_kg', 0.0) * user_data.get('mixed_beer_box_price', 0.0)
        oil_total = user_data.get('oil_kg', 0.0) * user_data.get('oil_price', 0.0)
        colored_plastic_total = user_data.get('colored_plastic_kg', 0.0) * user_data.get('colored_plastic_price', 0.0)
        iron_total = user_data.get('iron_kg', 0.0) * user_data.get('iron_price', 0.0)
        plastic_bag_total = user_data.get('plastic_bag_kg', 0.0) * user_data.get('plastic_bag_price', 0.0)
        mix_total = user_data.get('mix_kg', 0.0) * user_data.get('mix_price', 0.0)
        other_total = user_data.get('other_kg', 0.0) * user_data.get('other_price', 0.0)

        # Общая сумма
        total_pay = (
            pet_total + paper_total + alum_total + glass_total + 
            small_beer_box_total + large_beer_box_total + mixed_beer_box_total +
            oil_total + colored_plastic_total + iron_total + 
            plastic_bag_total + mix_total + other_total
        )

        # Добавляем отгрузку со всеми полями
        await add_shipment(
            point_id=user_data['point_id'],
            user_id=user.tg_id,
            pet_kg=user_data.get('pet_kg', 0.0),
            pet_price=user_data.get('pet_price', 0.0),
            pet_total=pet_total,
            paper_kg=user_data.get('paper_kg', 0.0),
            paper_price=user_data.get('paper_price', 0.0),
            paper_total=paper_total,
            alum_kg=user_data.get('alum_kg', 0.0),
            alum_price=user_data.get('alum_price', 0.0),
            alum_total=alum_total,
            glass_kg=user_data.get('glass_kg', 0.0),
            glass_price=user_data.get('glass_price', 0.0),
            glass_total=glass_total,
            small_beer_box_kg=user_data.get('small_beer_box_kg', 0.0),
            small_beer_box_price=user_data.get('small_beer_box_price', 0.0),
            small_beer_box_total=small_beer_box_total,
            large_beer_box_kg=user_data.get('large_beer_box_kg', 0.0),
            large_beer_box_price=user_data.get('large_beer_box_price', 0.0),
            large_beer_box_total=large_beer_box_total,
            mixed_beer_box_kg=user_data.get('mixed_beer_box_kg', 0.0),
            mixed_beer_box_price=user_data.get('mixed_beer_box_price', 0.0),
            mixed_beer_box_total=mixed_beer_box_total,
            oil_kg=user_data.get('oil_kg', 0.0),
            oil_price=user_data.get('oil_price', 0.0),
            oil_total=oil_total,
            colored_plastic_kg=user_data.get('colored_plastic_kg', 0.0),
            colored_plastic_price=user_data.get('colored_plastic_price', 0.0),
            colored_plastic_total=colored_plastic_total,
            iron_kg=user_data.get('iron_kg', 0.0),
            iron_price=user_data.get('iron_price', 0.0),
            iron_total=iron_total,
            plastic_bag_kg=user_data.get('plastic_bag_kg', 0.0),
            plastic_bag_price=user_data.get('plastic_bag_price', 0.0),
            plastic_bag_total=plastic_bag_total,
            mix_kg=user_data.get('mix_kg', 0.0),
            mix_price=user_data.get('mix_price', 0.0),
            mix_total=mix_total,
            other_kg=user_data.get('other_kg', 0.0),
            other_price=user_data.get('other_price', 0.0),
            other_total=other_total,
            total_pay=total_pay
        )
        
        # Очищаем количество мешков, так как отгрузка успешна
        await update_bags_count(user_data['point_id'], 0)
        
        # Отправляем уведомление пользователю, если он существует
        point = await get_point_by_id(user_data['point_id'])
        if point:
            point_user = await get_user_by_point_id(user_data['point_id'])
            if point_user:
                total_weight = (
                    user_data.get('pet_kg', 0) +
                    user_data.get('paper_kg', 0) +
                    user_data.get('alum_kg', 0) +
                    user_data.get('glass_kg', 0) +
                    user_data.get('small_beer_box_kg', 0) +
                    user_data.get('large_beer_box_kg', 0) +
                    user_data.get('mixed_beer_box_kg', 0) +
                    user_data.get('oil_kg', 0) +
                    user_data.get('colored_plastic_kg', 0) +
                    user_data.get('iron_kg', 0) +
                    user_data.get('plastic_bag_kg', 0) +
                    user_data.get('mix_kg', 0) +
                    user_data.get('other_kg', 0)
                )
                
                await callback.bot.send_message(
                    user.tg_id,
                    f"✅ การจัดส่งขยะของคุณได้รับการประมวลผลแล้ว\n\n"
                    f"📦 น้ำหนักรวม: {total_weight:.2f} กก.\n"
                    f"💰 ราคารวม: {total_pay:.2f} บาท\n\n"
                    f"ขอบคุณค่ะ/ครับ!"
                )
        
        await callback.message.edit_text(
            '✅ บันทึกข้อมูลการจัดส่งเรียบร้อย! จำนวนถุงถูกรีเซ็ตเป็นศูนย์',
            reply_markup=driver_keyboard()
        )
    except ValueError as e:
        await callback.message.edit_text(f"❌ ข้อผิดพลาด: {str(e)}")
    except Exception as e:
        await callback.message.edit_text(f"❌ เกิดข้อผิดพลาดที่ไม่คาดคิด: {str(e)}")
    finally:
        await state.clear()
        await callback.answer()

@admin.callback_query(F.data == "add_shipment")
async def process_add_shipment(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer('กรุณากรอก ID จุด:')  # "Введите ID точки:"
    await state.set_state(ShipmentStates.point_id)
    await callback.answer()

@admin.message(ShipmentStates.point_id)
async def process_point_id(message: Message, state: FSMContext):
    try:
        point_id = int(message.text)
        # Проверяем существование точки
        point = await get_point_by_id(point_id)
        if not point:
            await message.answer("ข้อผิดพลาด: ไม่พบจุดนี้ในระบบ กรุณากรอกใหม่")  # "Ошибка: Точка не найдена в системе. Пожалуйста, введите ID точки заново."
            return
            
        await state.update_data(point_id=point_id)
        await message.answer('กรุณาเลือกประเภท:', reply_markup=get_category_keyboard()) #Пожалуйста, выберите тип
    except ValueError:
        await message.answer("ข้อผิดพลาด: ID จุดต้องเป็นตัวเลขเต็ม กรุณากรอกใหม่") #Ошибка: идентификатор точки должен быть целым числом. пожалуйста, введите новый.

# Обработчики выбора категорий
@admin.callback_query(F.data == "category_main")
async def select_main_materials(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("กรุณาเลือกวัสดุ:", reply_markup=get_main_materials_keyboard())  # "Выберите материал:"
    await callback.answer()

@admin.callback_query(F.data == "category_other")
async def select_other_materials(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("กรุณาเลือกวัสดุ:", reply_markup=get_other_materials_keyboard())  # "Выберите материал:"
    await callback.answer()

@admin.callback_query(F.data == "back_to_categories")
async def back_to_categories(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("กรุณาเลือกประเภท:", reply_markup=get_category_keyboard())  # "Выберите категорию:"
    await callback.answer()

# Обработчики для основных материалов
@admin.callback_query(F.data == "material_pet")
async def process_pet_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Plastic PET (กก.):")  # "Введите вес Plastic PET (кг):"
    await state.set_state(ShipmentStates.pet_kg)
    await callback.answer()

@admin.message(ShipmentStates.pet_kg)
async def process_pet_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        pet_kg = float(text)
        if pet_kg < 0:
            raise ValueError
        await state.update_data(pet_kg=pet_kg)
        
        if pet_kg == 0:
            await state.update_data(pet_price=0.0)
            await message.answer("น้ำหนัก Plastic PET เป็น 0", reply_markup=get_category_keyboard())  # "Вес Plastic PET равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Plastic PET:")  # "Введите цену за кг Plastic PET:"
            await state.set_state(ShipmentStates.pet_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.pet_price)
async def process_pet_price(message: Message, state: FSMContext):
    try:
        pet_price = float(message.text)
        if pet_price < 0:
            raise ValueError
        await state.update_data(pet_price=pet_price)
        await message.answer("บันทึกข้อมูล Plastic PET เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Plastic PET сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Paper
@admin.callback_query(F.data == "material_paper")
async def process_paper_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Paper (กก.):")  # "Введите вес Paper (кг):"
    await state.set_state(ShipmentStates.paper_kg)
    await callback.answer()

@admin.message(ShipmentStates.paper_kg)
async def process_paper_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        paper_kg = float(text)
        if paper_kg < 0:
            raise ValueError
        await state.update_data(paper_kg=paper_kg)
        
        if paper_kg == 0:
            await state.update_data(paper_price=0.0)
            await message.answer("น้ำหนัก Paper เป็น 0", reply_markup=get_category_keyboard())  # "Вес Paper равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Paper:")  # "Введите цену за кг Paper:"
            await state.set_state(ShipmentStates.paper_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.paper_price)
async def process_paper_price(message: Message, state: FSMContext):
    try:
        paper_price = float(message.text)
        if paper_price < 0:
            raise ValueError
        await state.update_data(paper_price=paper_price)
        await message.answer("บันทึกข้อมูล Paper เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Paper сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Aluminum
@admin.callback_query(F.data == "material_alum")
async def process_alum_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Aluminum (กก.):")  # "Введите вес Aluminum (кг):"
    await state.set_state(ShipmentStates.alum_kg)
    await callback.answer()

@admin.message(ShipmentStates.alum_kg)
async def process_alum_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        alum_kg = float(text)
        if alum_kg < 0:
            raise ValueError
        await state.update_data(alum_kg=alum_kg)
        
        if alum_kg == 0:
            await state.update_data(alum_price=0.0)
            await message.answer("น้ำหนัก Aluminum เป็น 0", reply_markup=get_category_keyboard())  # "Вес Aluminum равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Aluminum:")  # "Введите цену за кг Aluminum:"
            await state.set_state(ShipmentStates.alum_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.alum_price)
async def process_alum_price(message: Message, state: FSMContext):
    try:
        alum_price = float(message.text)
        if alum_price < 0:
            raise ValueError
        await state.update_data(alum_price=alum_price)
        await message.answer("บันทึกข้อมูล Aluminum เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Aluminum сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Glass
@admin.callback_query(F.data == "material_glass")
async def process_glass_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Glass (กก.):")  # "Введите вес Glass (кг):"
    await state.set_state(ShipmentStates.glass_kg)
    await callback.answer()

@admin.message(ShipmentStates.glass_kg)
async def process_glass_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        glass_kg = float(text)
        if glass_kg < 0:
            raise ValueError
        await state.update_data(glass_kg=glass_kg)
        
        if glass_kg == 0:
            await state.update_data(glass_price=0.0)
            await message.answer("น้ำหนัก Glass เป็น 0", reply_markup=get_category_keyboard())  # "Вес Glass равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Glass:")  # "Введите цену за кг Glass:"
            await state.set_state(ShipmentStates.glass_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.glass_price)
async def process_glass_price(message: Message, state: FSMContext):
    try:
        glass_price = float(message.text)
        if glass_price < 0:
            raise ValueError
        await state.update_data(glass_price=glass_price)
        await message.answer("บันทึกข้อมูล Glass เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Glass сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Small Beer Box
@admin.callback_query(F.data == "material_small_beer_box")
async def process_small_beer_box_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Small Beer Box (กก.):")
    await state.set_state(ShipmentStates.small_beer_box_kg)
    await callback.answer()

@admin.message(ShipmentStates.small_beer_box_kg)
async def process_small_beer_box_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        small_beer_box_kg = float(text)
        if small_beer_box_kg < 0:
            raise ValueError
        await state.update_data(small_beer_box_kg=small_beer_box_kg)
        
        if small_beer_box_kg == 0:
            await state.update_data(small_beer_box_price=0.0)
            await message.answer("น้ำหนัก Small Beer Box เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Small Beer Box:")
            await state.set_state(ShipmentStates.small_beer_box_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.small_beer_box_price)
async def process_small_beer_box_price(message: Message, state: FSMContext):
    try:
        small_beer_box_price = float(message.text)
        if small_beer_box_price < 0:
            raise ValueError
        await state.update_data(small_beer_box_price=small_beer_box_price)
        await message.answer("บันทึกข้อมูล Small Beer Box เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Large Beer Box
@admin.callback_query(F.data == "material_large_beer_box")
async def process_large_beer_box_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Large Beer Box (กก.):")
    await state.set_state(ShipmentStates.large_beer_box_kg)
    await callback.answer()

@admin.message(ShipmentStates.large_beer_box_kg)
async def process_large_beer_box_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        large_beer_box_kg = float(text)
        if large_beer_box_kg < 0:
            raise ValueError
        await state.update_data(large_beer_box_kg=large_beer_box_kg)
        
        if large_beer_box_kg == 0:
            await state.update_data(large_beer_box_price=0.0)
            await message.answer("น้ำหนัก Large Beer Box เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Large Beer Box:")
            await state.set_state(ShipmentStates.large_beer_box_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.large_beer_box_price)
async def process_large_beer_box_price(message: Message, state: FSMContext):
    try:
        large_beer_box_price = float(message.text)
        if large_beer_box_price < 0:
            raise ValueError
        await state.update_data(large_beer_box_price=large_beer_box_price)
        await message.answer("บันทึกข้อมูล Large Beer Box เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Mixed Beer Box
@admin.callback_query(F.data == "material_mixed_beer_box")
async def process_mixed_beer_box_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Mixed Beer Box (กก.):")
    await state.set_state(ShipmentStates.mixed_beer_box_kg)
    await callback.answer()

@admin.message(ShipmentStates.mixed_beer_box_kg)
async def process_mixed_beer_box_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        mixed_beer_box_kg = float(text)
        if mixed_beer_box_kg < 0:
            raise ValueError
        await state.update_data(mixed_beer_box_kg=mixed_beer_box_kg)
        
        if mixed_beer_box_kg == 0:
            await state.update_data(mixed_beer_box_price=0.0)
            await message.answer("น้ำหนัก Mixed Beer Box เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Mixed Beer Box:")
            await state.set_state(ShipmentStates.mixed_beer_box_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.mixed_beer_box_price)
async def process_mixed_beer_box_price(message: Message, state: FSMContext):
    try:
        mixed_beer_box_price = float(message.text)
        if mixed_beer_box_price < 0:
            raise ValueError
        await state.update_data(mixed_beer_box_price=mixed_beer_box_price)
        await message.answer("บันทึกข้อมูล Mixed Beer Box เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для других материалов (категория 2)
# Обработчики для Oil
@admin.callback_query(F.data == "material_oil")
async def process_oil_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Oil (กก.):")  # "Введите вес Oil (кг):"
    await state.set_state(ShipmentStates.oil_kg)
    await callback.answer()

@admin.message(ShipmentStates.oil_kg)
async def process_oil_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        oil_kg = float(text)
        if oil_kg < 0:
            raise ValueError
        await state.update_data(oil_kg=oil_kg)
        
        if oil_kg == 0:
            await state.update_data(oil_price=0.0)
            await message.answer("น้ำหนัก Oil เป็น 0", reply_markup=get_category_keyboard())  # "Вес Oil равен 0."
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Oil:")  # "Введите цену за кг Oil:"
            await state.set_state(ShipmentStates.oil_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Вес должен быть положительным числом. Введите вес заново:"

@admin.message(ShipmentStates.oil_price)
async def process_oil_price(message: Message, state: FSMContext):
    try:
        oil_price = float(message.text)
        if oil_price < 0:
            raise ValueError
        await state.update_data(oil_price=oil_price)
        await message.answer("บันทึกข้อมูล Oil เรียบร้อย", reply_markup=get_category_keyboard())  # "Данные по Oil сохранены."
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")  # "Ошибка: Цена должна быть положительным числом. Введите цену заново:"

# Обработчики для Colored Plastic
@admin.callback_query(F.data == "material_colored_plastic")
async def process_colored_plastic_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Colored Plastic (กก.):")
    await state.set_state(ShipmentStates.colored_plastic_kg)
    await callback.answer()

@admin.message(ShipmentStates.colored_plastic_kg)
async def process_colored_plastic_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        colored_plastic_kg = float(text)
        if colored_plastic_kg < 0:
            raise ValueError
        await state.update_data(colored_plastic_kg=colored_plastic_kg)
        
        if colored_plastic_kg == 0:
            await state.update_data(colored_plastic_price=0.0)
            await message.answer("น้ำหนัก Colored Plastic เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Colored Plastic:")
            await state.set_state(ShipmentStates.colored_plastic_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.colored_plastic_price)
async def process_colored_plastic_price(message: Message, state: FSMContext):
    try:
        colored_plastic_price = float(message.text)
        if colored_plastic_price < 0:
            raise ValueError
        await state.update_data(colored_plastic_price=colored_plastic_price)
        await message.answer("บันทึกข้อมูล Colored Plastic เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Iron
@admin.callback_query(F.data == "material_iron")
async def process_iron_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Iron (กก.):")
    await state.set_state(ShipmentStates.iron_kg)
    await callback.answer()

@admin.message(ShipmentStates.iron_kg)
async def process_iron_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        iron_kg = float(text)
        if iron_kg < 0:
            raise ValueError
        await state.update_data(iron_kg=iron_kg)
        
        if iron_kg == 0:
            await state.update_data(iron_price=0.0)
            await message.answer("น้ำหนัก Iron เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Iron:")
            await state.set_state(ShipmentStates.iron_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.iron_price)
async def process_iron_price(message: Message, state: FSMContext):
    try:
        iron_price = float(message.text)
        if iron_price < 0:
            raise ValueError
        await state.update_data(iron_price=iron_price)
        await message.answer("บันทึกข้อมูล Iron เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Plastic Bag or Container
@admin.callback_query(F.data == "material_plastic_bag")
async def process_plastic_bag_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Plastic Bag or Container (กก.):")
    await state.set_state(ShipmentStates.plastic_bag_kg)
    await callback.answer()

@admin.message(ShipmentStates.plastic_bag_kg)
async def process_plastic_bag_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        plastic_bag_kg = float(text)
        if plastic_bag_kg < 0:
            raise ValueError
        await state.update_data(plastic_bag_kg=plastic_bag_kg)
        
        if plastic_bag_kg == 0:
            await state.update_data(plastic_bag_price=0.0)
            await message.answer("น้ำหนัก Plastic Bag or Container เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Plastic Bag or Container:")
            await state.set_state(ShipmentStates.plastic_bag_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.plastic_bag_price)
async def process_plastic_bag_price(message: Message, state: FSMContext):
    try:
        plastic_bag_price = float(message.text)
        if plastic_bag_price < 0:
            raise ValueError
        await state.update_data(plastic_bag_price=plastic_bag_price)
        await message.answer("บันทึกข้อมูล Plastic Bag or Container เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Mix
@admin.callback_query(F.data == "material_mix")
async def process_mix_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Mix (กก.):")
    await state.set_state(ShipmentStates.mix_kg)
    await callback.answer()

@admin.message(ShipmentStates.mix_kg)
async def process_mix_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        mix_kg = float(text)
        if mix_kg < 0:
            raise ValueError
        await state.update_data(mix_kg=mix_kg)
        
        if mix_kg == 0:
            await state.update_data(mix_price=0.0)
            await message.answer("น้ำหนัก Mix เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Mix:")
            await state.set_state(ShipmentStates.mix_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.mix_price)
async def process_mix_price(message: Message, state: FSMContext):
    try:
        mix_price = float(message.text)
        if mix_price < 0:
            raise ValueError
        await state.update_data(mix_price=mix_price)
        await message.answer("บันทึกข้อมูล Mix เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчики для Other
@admin.callback_query(F.data == "material_other")
async def process_other_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("กรุณากรอกน้ำหนัก Other (กก.):")
    await state.set_state(ShipmentStates.other_kg)
    await callback.answer()

@admin.message(ShipmentStates.other_kg)
async def process_other_kg(message: Message, state: FSMContext):
    try:
        text = message.text.replace(',', '.').strip()
        other_kg = float(text)
        if other_kg < 0:
            raise ValueError
        await state.update_data(other_kg=other_kg)
        
        if other_kg == 0:
            await state.update_data(other_price=0.0)
            await message.answer("น้ำหนัก Other เป็น 0", reply_markup=get_category_keyboard())
        else:
            await message.answer("กรุณากรอกราคาต่อกก. Other:")
            await state.set_state(ShipmentStates.other_price)
    except ValueError:
        await message.answer("ข้อผิดพลาด: น้ำหนักต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

@admin.message(ShipmentStates.other_price)
async def process_other_price(message: Message, state: FSMContext):
    try:
        other_price = float(message.text)
        if other_price < 0:
            raise ValueError
        await state.update_data(other_price=other_price)
        await message.answer("บันทึกข้อมูล Other เรียบร้อย", reply_markup=get_category_keyboard())
    except ValueError:
        await message.answer("ข้อผิดพลาด: ราคาต้องเป็นตัวเลขบวก กรุณากรอกใหม่")

# Обработчик завершения ввода
@admin.callback_query(F.data == "finish_shipment")
async def finish_shipment(callback: CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    
    # Формируем сводку по всем введенным данным
    summary = "📋 สรุปการจัดส่ง:\n\n"  # "📋 Сводка по отгрузке:"
    total_weight = 0
    total_cost = 0
    
    # Основные материалы
    if 'pet_kg' in user_data:
        pet_cost = user_data.get('pet_kg', 0) * user_data.get('pet_price', 0)
        summary += f"🔹 Plastic PET: {user_data['pet_kg']} กก., {pet_cost:.2f} บาท\n"
        total_weight += user_data['pet_kg']
        total_cost += pet_cost
    
    if 'paper_kg' in user_data:
        paper_cost = user_data.get('paper_kg', 0) * user_data.get('paper_price', 0)
        summary += f"🔹 Paper: {user_data['paper_kg']} กก., {paper_cost:.2f} บาท\n"
        total_weight += user_data['paper_kg']
        total_cost += paper_cost
    
    if 'alum_kg' in user_data:
        alum_cost = user_data.get('alum_kg', 0) * user_data.get('alum_price', 0)
        summary += f"🔹 Aluminum: {user_data['alum_kg']} กก., {alum_cost:.2f} บาท\n"
        total_weight += user_data['alum_kg']
        total_cost += alum_cost
    
    if 'glass_kg' in user_data:
        glass_cost = user_data.get('glass_kg', 0) * user_data.get('glass_price', 0)
        summary += f"🔹 Glass: {user_data['glass_kg']} กก., {glass_cost:.2f} บาท\n"
        total_weight += user_data['glass_kg']
        total_cost += glass_cost
    
    if 'small_beer_box_kg' in user_data:
        small_beer_box_cost = user_data.get('small_beer_box_kg', 0) * user_data.get('small_beer_box_price', 0)
        summary += f"🔹 Small Beer Box: {user_data['small_beer_box_kg']} กก., {small_beer_box_cost:.2f} บาท\n"
        total_weight += user_data['small_beer_box_kg']
        total_cost += small_beer_box_cost
    
    if 'large_beer_box_kg' in user_data:
        large_beer_box_cost = user_data.get('large_beer_box_kg', 0) * user_data.get('large_beer_box_price', 0)
        summary += f"🔹 Large Beer Box: {user_data['large_beer_box_kg']} กก., {large_beer_box_cost:.2f} บาท\n"
        total_weight += user_data['large_beer_box_kg']
        total_cost += large_beer_box_cost
    
    if 'mixed_beer_box_kg' in user_data:
        mixed_beer_box_cost = user_data.get('mixed_beer_box_kg', 0) * user_data.get('mixed_beer_box_price', 0)
        summary += f"🔹 Mixed Beer Box: {user_data['mixed_beer_box_kg']} กก., {mixed_beer_box_cost:.2f} บาท\n"
        total_weight += user_data['mixed_beer_box_kg']
        total_cost += mixed_beer_box_cost
    
    # Другие материалы
    if 'oil_kg' in user_data:
        oil_cost = user_data.get('oil_kg', 0) * user_data.get('oil_price', 0)
        summary += f"🔸 Oil: {user_data['oil_kg']} กก., {oil_cost:.2f} บาท\n"
        total_weight += user_data['oil_kg']
        total_cost += oil_cost
    
    if 'colored_plastic_kg' in user_data:
        colored_plastic_cost = user_data.get('colored_plastic_kg', 0) * user_data.get('colored_plastic_price', 0)
        summary += f"🔸 Colored Plastic: {user_data['colored_plastic_kg']} กก., {colored_plastic_cost:.2f} บาท\n"
        total_weight += user_data['colored_plastic_kg']
        total_cost += colored_plastic_cost
    
    if 'iron_kg' in user_data:
        iron_cost = user_data.get('iron_kg', 0) * user_data.get('iron_price', 0)
        summary += f"🔸 Iron: {user_data['iron_kg']} กก., {iron_cost:.2f} บาท\n"
        total_weight += user_data['iron_kg']
        total_cost += iron_cost
    
    if 'plastic_bag_kg' in user_data:
        plastic_bag_cost = user_data.get('plastic_bag_kg', 0) * user_data.get('plastic_bag_price', 0)
        summary += f"🔸 Plastic Bag or Container: {user_data['plastic_bag_kg']} กก., {plastic_bag_cost:.2f} บาท\n"
        total_weight += user_data['plastic_bag_kg']
        total_cost += plastic_bag_cost
    
    if 'mix_kg' in user_data:
        mix_cost = user_data.get('mix_kg', 0) * user_data.get('mix_price', 0)
        summary += f"🔸 Mix: {user_data['mix_kg']} กก., {mix_cost:.2f} บาท\n"
        total_weight += user_data['mix_kg']
        total_cost += mix_cost
    
    if 'other_kg' in user_data:
        other_cost = user_data.get('other_kg', 0) * user_data.get('other_price', 0)
        summary += f"🔸 Other: {user_data['other_kg']} กก., {other_cost:.2f} บาท\n"
        total_weight += user_data['other_kg']
        total_cost += other_cost
    
    summary += f"\n💎 รวม: {total_weight:.2f} กก., {total_cost:.2f} บาท"  # "💎 Итого: кг, руб."
    
    await callback.message.edit_text(summary, reply_markup=get_confirmation_keyboard())
    await callback.answer()

# Обработчик подтверждения
@admin.callback_query(F.data == "confirm_shipment")
async def confirm_shipment(callback: CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    
    try:
        user = await get_user_by_tg_id(callback.from_user.id)
        if not user:
            await callback.answer()  # Просто закрываем callback без сообщения
            return  # Выходим из функции, если пользователь не найден
        
        # Рассчитываем общие суммы для каждого материала
        pet_total = user_data.get('pet_kg', 0.0) * user_data.get('pet_price', 0.0)
        paper_total = user_data.get('paper_kg', 0.0) * user_data.get('paper_price', 0.0)
        alum_total = user_data.get('alum_kg', 0.0) * user_data.get('alum_price', 0.0)
        glass_total = user_data.get('glass_kg', 0.0) * user_data.get('glass_price', 0.0)
        small_beer_box_total = user_data.get('small_beer_box_kg', 0.0) * user_data.get('small_beer_box_price', 0.0)
        large_beer_box_total = user_data.get('large_beer_box_kg', 0.0) * user_data.get('large_beer_box_price', 0.0)
        mixed_beer_box_total = user_data.get('mixed_beer_box_kg', 0.0) * user_data.get('mixed_beer_box_price', 0.0)
        oil_total = user_data.get('oil_kg', 0.0) * user_data.get('oil_price', 0.0)
        colored_plastic_total = user_data.get('colored_plastic_kg', 0.0) * user_data.get('colored_plastic_price', 0.0)
        iron_total = user_data.get('iron_kg', 0.0) * user_data.get('iron_price', 0.0)
        plastic_bag_total = user_data.get('plastic_bag_kg', 0.0) * user_data.get('plastic_bag_price', 0.0)
        mix_total = user_data.get('mix_kg', 0.0) * user_data.get('mix_price', 0.0)
        other_total = user_data.get('other_kg', 0.0) * user_data.get('other_price', 0.0)

        # Общая сумма
        total_pay = (
            pet_total + paper_total + alum_total + glass_total + 
            small_beer_box_total + large_beer_box_total + mixed_beer_box_total +
            oil_total + colored_plastic_total + iron_total + 
            plastic_bag_total + mix_total + other_total
        )

        # Добавляем отгрузку со всеми полями
        await add_shipment(
            point_id=user_data['point_id'],
            user_id=user.tg_id,
            pet_kg=user_data.get('pet_kg', 0.0),
            pet_price=user_data.get('pet_price', 0.0),
            pet_total=pet_total,
            paper_kg=user_data.get('paper_kg', 0.0),
            paper_price=user_data.get('paper_price', 0.0),
            paper_total=paper_total,
            alum_kg=user_data.get('alum_kg', 0.0),
            alum_price=user_data.get('alum_price', 0.0),
            alum_total=alum_total,
            glass_kg=user_data.get('glass_kg', 0.0),
            glass_price=user_data.get('glass_price', 0.0),
            glass_total=glass_total,
            small_beer_box_kg=user_data.get('small_beer_box_kg', 0.0),
            small_beer_box_price=user_data.get('small_beer_box_price', 0.0),
            small_beer_box_total=small_beer_box_total,
            large_beer_box_kg=user_data.get('large_beer_box_kg', 0.0),
            large_beer_box_price=user_data.get('large_beer_box_price', 0.0),
            large_beer_box_total=large_beer_box_total,
            mixed_beer_box_kg=user_data.get('mixed_beer_box_kg', 0.0),
            mixed_beer_box_price=user_data.get('mixed_beer_box_price', 0.0),
            mixed_beer_box_total=mixed_beer_box_total,
            oil_kg=user_data.get('oil_kg', 0.0),
            oil_price=user_data.get('oil_price', 0.0),
            oil_total=oil_total,
            colored_plastic_kg=user_data.get('colored_plastic_kg', 0.0),
            colored_plastic_price=user_data.get('colored_plastic_price', 0.0),
            colored_plastic_total=colored_plastic_total,
            iron_kg=user_data.get('iron_kg', 0.0),
            iron_price=user_data.get('iron_price', 0.0),
            iron_total=iron_total,
            plastic_bag_kg=user_data.get('plastic_bag_kg', 0.0),
            plastic_bag_price=user_data.get('plastic_bag_price', 0.0),
            plastic_bag_total=plastic_bag_total,
            mix_kg=user_data.get('mix_kg', 0.0),
            mix_price=user_data.get('mix_price', 0.0),
            mix_total=mix_total,
            other_kg=user_data.get('other_kg', 0.0),
            other_price=user_data.get('other_price', 0.0),
            other_total=other_total,
            total_pay=total_pay
        )
        
        # Очищаем количество мешков, так как отгрузка успешна
        await update_bags_count(user_data['point_id'], 0)
        
        # Отправляем уведомление пользователю, если он существует
        point = await get_point_by_id(user_data['point_id'])
        if point:
            point_user = await get_user_by_point_id(user_data['point_id'])
            if point_user:
                total_weight = (
                    user_data.get('pet_kg', 0) +
                    user_data.get('paper_kg', 0) +
                    user_data.get('alum_kg', 0) +
                    user_data.get('glass_kg', 0) +
                    user_data.get('small_beer_box_kg', 0) +
                    user_data.get('large_beer_box_kg', 0) +
                    user_data.get('mixed_beer_box_kg', 0) +
                    user_data.get('oil_kg', 0) +
                    user_data.get('colored_plastic_kg', 0) +
                    user_data.get('iron_kg', 0) +
                    user_data.get('plastic_bag_kg', 0) +
                    user_data.get('mix_kg', 0) +
                    user_data.get('other_kg', 0)
                )
                
                await callback.bot.send_message(
                    user.tg_id,
                    f"✅ การจัดส่งขยะของคุณได้รับการประมวลผลแล้ว\n\n"
                    f"📦 น้ำหนักรวม: {total_weight:.2f} กก.\n"
                    f"💰 ราคารวม: {total_pay:.2f} บาท\n\n"
                    f"ขอบคุณค่ะ/ครับ!"
                )
        
        await callback.message.edit_text(
            '✅ บันทึกข้อมูลการจัดส่งเรียบร้อย! จำนวนถุงถูกรีเซ็ตเป็นศูนย์',  # "Данные об отгрузке успешно добавлены! Количество мешков обнулено."
            reply_markup=driver_keyboard()
        )
    except ValueError as e:
        await callback.message.edit_text(f"❌ ข้อผิดพลาด: {str(e)}")
    except Exception as e:
        await callback.message.edit_text(f"❌ เกิดข้อผิดพลาดที่ไม่คาดคิด: {str(e)}")
    finally:
        await state.clear()
        await callback.answer()

# Обработчик отмены
@admin.callback_query(F.data == "cancel_shipment")
async def cancel_shipment(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "❌ ยกเลิกการจัดส่ง ข้อมูลทั้งหมดถูกลบแล้ว",  # "❌ Отгрузка отменена. Все данные удалены."
    )
    await callback.answer()

# Обработчик для кнопки "Отменить" в процессе ввода
@admin.callback_query(F.data == "cancel_during_input")
async def cancel_during_input(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "❌ ยกเลิกการป้อนข้อมูล ข้อมูลชั่วคราวทั้งหมดถูกลบ",  # "❌ Ввод данных прерван. Все временные данные удалены."
    )
    await callback.answer()


@admin.callback_query(F.data == "create_point")
async def start_create_point(callback: CallbackQuery, state: FSMContext):
    """Начало процесса создания точки"""
    await callback.message.answer(
        "กรุณากรอก ID จุดใหม่ในรูปแบบ RZZN โดยที่:\n"  # "Введите ID новой точки в формате RZZN, где:"
        "R - หมายเลขภูมิภาค (1 หลัก)\n"  # "R - номер региона (1 цифра)"
        "ZZ - หมายเลขโซน (2 หลัก)\n"  # "ZZ - номер зоны (2 цифры)"
        "N - หมายเลขจุดในโซน (1 หลัก)\n"  # "N - номер точки в зоне (1 цифра)"
        "ตัวอย่าง: 1021 - จุดในภูมิภาค 1 โซน 02 จุดที่ 1",  # "Пример: 1021 - точка в регионе 1, зоне 02, номер точки 1"
        reply_markup=cancel_keyboard()
    )
    await state.set_state(CreatePoint.point_id)
    await callback.answer()

@admin.message(CreatePoint.point_id, F.text.regexp(r'^\d{4}$'))
async def process_point_id(message: Message, state: FSMContext):
    """Обработка ID точки с проверкой формата"""
    point_id = message.text
    region_id = int(point_id[0])
    zone_num = int(point_id[1:3])
    zone_id = int(f"{region_id}{zone_num:02d}")
    point_num = int(point_id[3])
    
    # Проверяем, существует ли уже точка с таким ID
    if await get_point_by_id(point_id):
        await message.answer(
            "มีจุดนี้อยู่แล้ว! กรุณากรอก ID อื่น",  # "Точка с таким ID уже существует! Пожалуйста, введите другой ID."
            reply_markup=cancel_keyboard()
        )
        return
    
    await state.update_data(
        point_id=point_id,
        region_id=region_id,
        zone_num=zone_num,
        zone_id=zone_id,
        point_num=point_num
    )
    
    await message.answer(
        f"ID จุด: {point_id}\n"
        f"ภูมิภาค: {region_id}\n"
        f"หมายเลขโซน: {zone_num}\n"
        f"หมายเลขจุด: {point_num}\n\n"
        "กรุณากรอกชื่อจุด:",  # "Введите название точки:"
        reply_markup=cancel_keyboard()
    )
    await state.set_state(CreatePoint.point_name)

@admin.message(CreatePoint.point_id)
async def process_point_id_invalid(message: Message):
    """Обработка неверного формата ID точки"""
    await message.answer(
        "รูปแบบ ID จุดไม่ถูกต้อง! ต้องเป็นตัวเลข 4 หลักในรูปแบบ RZZN\n"
        "ตัวอย่าง: 1021 - จุดในภูมิภาค 1 โซน 02 จุดที่ 1",  # "Неверный формат ID точки! Должно быть 4 цифры в формате RZZN"
        reply_markup=cancel_keyboard()
    )

@admin.message(CreatePoint.point_name)
async def process_point_name(message: Message, state: FSMContext):
    """Обработка названия точки"""
    await state.update_data(point_name=message.text)
    await message.answer(
        "กรุณากรอกชื่อเจ้าของจุด:",  # "Введите имя владельца точки:"
        reply_markup=cancel_keyboard()
    )
    await state.set_state(CreatePoint.point_owner_name)

@admin.message(CreatePoint.point_owner_name)
async def process_owner_name(message: Message, state: FSMContext):
    """Обработка имени владельца"""
    await state.update_data(point_owner_name=message.text)
    await message.answer(
        "กรุณากรอกเบอร์โทรเจ้าของจุด:",  # "Введите номер телефона владельца:"
        reply_markup=cancel_keyboard()
    )
    await state.set_state(CreatePoint.phone_number)

@admin.message(CreatePoint.phone_number)
async def process_phone_number(message: Message, state: FSMContext):
    """Обработка номера телефона"""
    await state.update_data(phone_number=message.text)
    await message.answer(
        "กรุณากรอกที่อยู่จุด:",  # "Введите адрес точки:"
        reply_markup=cancel_keyboard()
    )
    await state.set_state(CreatePoint.address)

@admin.message(CreatePoint.address)
async def process_address(message: Message, state: FSMContext):
    """Обработка адреса"""
    await state.update_data(address=message.text)
    
    data = await state.get_data()
    
    # Формируем сообщение для подтверждения
    confirm_text = (
        "กรุณาตรวจสอบข้อมูลที่ป้อน:\n\n"  # "Проверьте введенные данные:"
        f"ID จุด: {data['point_id']}\n"
        f"ภูมิภาค: {data['region_id']}\n"
        f"โซน: {data['zone_num']}\n"
        f"หมายเลขจุด: {data['point_num']}\n"
        f"ชื่อ: {data['point_name']}\n"
        f"เจ้าของ: {data['point_owner_name']}\n"
        f"โทรศัพท์: {data['phone_number']}\n"
        f"ที่อยู่: {data['address']}\n\n"
        "ถูกต้องทั้งหมดหรือไม่?"  # "Все верно?"
    )
    
    await message.answer(confirm_text, reply_markup=confirm_keyboard())
    await state.set_state(CreatePoint.confirmation)

@admin.callback_query(CreatePoint.confirmation, F.data == "confirm")
async def confirm_point_creation(callback: CallbackQuery, state: FSMContext):
    """Подтверждение создания точки"""
    data = await state.get_data()
    
    try:
        # Сначала проверяем/создаем регион
        if not await get_region_by_id(data['region_id']):
            await add_region(data['region_id'])
        
        # Затем проверяем/создаем зону (используем zone_id из 3 цифр)
        if not await get_zone_by_id(data['zone_id']):
            await add_zone(data['zone_id'], data['region_id'])
        
        # Создаем точку (используем point_id как строку)
        await add_point(
            point_id=data['point_id'],
            point_name=data['point_name'],
            point_owner_name=data['point_owner_name'],
            phone_number=data['phone_number'],
            address=data['address'],
            bags_count=0,
            zone_id=data['zone_id']
        )
        
        await callback.message.answer(
            "✅ สร้างจุดเรียบร้อยแล้ว!",  # "✅ Точка успешно создана!"
            reply_markup=admin_keyboard()
        )
    except Exception as e:
        await callback.message.answer(
            f"❌ ข้อผิดพลาดในการสร้างจุด: {str(e)}",  # "❌ Ошибка при создании точки:"
            reply_markup=admin_keyboard()
        )
    finally:
        await state.clear()

@admin.callback_query(CreatePoint.confirmation, F.data == "cancel")
async def cancel_point_creation(callback: CallbackQuery, state: FSMContext):
    """Отмена создания точки"""
    await callback.message.answer(
        "❌ ยกเลิกการสร้างจุด",  # "❌ Создание точки отменено."
        reply_markup=admin_keyboard()
    )
    await state.clear()

@admin.callback_query(StateFilter(CreatePoint), F.data == "cancel_operation")
async def cancel_creation_process(callback: CallbackQuery, state: FSMContext):
    """Отмена процесса создания на любом этапе"""
    await callback.message.answer(
        "❌ ยกเลิกกระบวนการสร้างจุด",  # "❌ Создание точки прервано."
        reply_markup=admin_keyboard()
    )
    await state.clear()



@admin.callback_query(F.data == "delete_point")
async def start_delete_point(callback: CallbackQuery, state: FSMContext):
    """Начало процесса удаления точки"""
    await callback.message.answer(
        "กรุณากรอก ID จุดที่ต้องการลบ:",  # "Введите ID точки для удаления:"
        reply_markup=cancel_keyboard()
    )
    await state.set_state("delete_point")
    await callback.answer()

@admin.message(F.text.regexp(r'^\d{4}$'), StateFilter("delete_point"))
async def process_delete_point_id(message: Message, state: FSMContext):
    """Обработка ID точки для удаления"""
    point_id = message.text
    
    # Проверяем существование точки
    point = await get_point_by_id(point_id)
    if not point:
        await message.answer(
            "ไม่พบจุดนี้ในระบบ! กรุณาตรวจสอบ ID อีกครั้ง",  # "Точка не найдена! Проверьте ID"
            reply_markup=cancel_keyboard()
        )
        return
    
    # Получаем пользователя, привязанного к точке
    user = await get_user_by_point_id(point_id)
    
    confirm_text = (
        "⚠️ คุณแน่ใจหรือไม่ว่าต้องการลบจุดนี้?\n\n"
        f"ID จุด: {point_id}\n"
    )
    
    if user:
        confirm_text += f"ผู้ใช้ที่ผูกติด: {user.tg_id}\n"
    
    confirm_text += (
        "\nการดำเนินการนี้จะ:\n"
        "- ลบจุดออกจากระบบ\n"
        "- ลบผู้ใช้ที่ผูกติด (ถ้ามี)\n"
        "- ลบคำขอทั้งหมดของจุดนี้\n"
        "- เก็บประวัติการจัดส่งไว้\n"
        "\nยืนยันการลบหรือไม่?"  # "Подтвердите удаление?"
    )
    
    await state.update_data(point_id=point_id)
    await message.answer(confirm_text, reply_markup=confirm_keyboard())
    await state.set_state("confirm_delete_point")

@admin.message(StateFilter("delete_point"))
async def process_delete_point_id_invalid(message: Message):
    """Обработка неверного формата ID точки при удалении"""
    await message.answer(
        "รูปแบบ ID จุดไม่ถูกต้อง! ต้องเป็นตัวเลข 4 หลัก\n"
        "ตัวอย่าง: 1021",
        reply_markup=cancel_keyboard()
    )

@admin.callback_query(StateFilter("confirm_delete_point"), F.data == "confirm")
async def confirm_point_deletion(callback: CallbackQuery, state: FSMContext):
    """Подтверждение удаления точки"""
    data = await state.get_data()
    point_id = str(data['point_id'])  # Убедимся, что передаем строку
    
    if await delete_point_and_related_data(point_id):
        await callback.message.answer(
            f"✅ จุด {point_id} ถูกลบออกจากระบบเรียบร้อย!",
            reply_markup=admin_keyboard()
        )
    else:
        await callback.message.answer(
            f"❌ เกิดข้อผิดพลาดในการลบจุด {point_id}",
            reply_markup=admin_keyboard()
        )
    
    await state.clear()

@admin.callback_query(StateFilter("confirm_delete_point"), F.data == "cancel")
async def cancel_point_deletion(callback: CallbackQuery, state: FSMContext):
    """Отмена удаления точки"""
    await callback.message.answer(
        "❌ ยกเลิกการลบจุด",  # "Удаление точки отменено"
        reply_markup=admin_keyboard()
    )
    await state.clear()